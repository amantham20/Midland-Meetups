#!/usr/bin/env python3
"""
Import Midland Meetups spreadsheet → Firestore.

Reads: conversions/Midland Meetups Database.xlsx
Writes collections: events, memories, squad, rsvps

Skips legacy-only tabs (Chat, Scores, WalterProgress) — not used by the rewrite.

Requirements:
  pip install openpyxl firebase-admin

Auth (one of):
  export FIREBASE_SERVICE_ACCOUNT_JSON='{"type":"service_account",...}'
  export GOOGLE_APPLICATION_CREDENTIALS=/path/to/serviceAccount.json

Usage:
  python3 conversions/import_to_firestore.py --dry-run
  python3 conversions/import_to_firestore.py
  python3 conversions/import_to_firestore.py --xlsx path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl:  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent
DEFAULT_XLSX = ROOT / "Midland Meetups Database.xlsx"


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def cell_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    s = cell_str(v).upper()
    return s in ("TRUE", "1", "YES", "Y")


def slug_name(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s or "guest"


def rows_as_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_str(h) or f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for raw in rows[1:]:
        if all(c is None or cell_str(c) == "" for c in raw):
            continue
        d = {}
        for i, h in enumerate(headers):
            if h.startswith("col") and h[3:].isdigit():
                continue
            d[h] = raw[i] if i < len(raw) else None
        out.append(d)
    return out


def map_event(row: dict) -> tuple[str, dict]:
    eid = cell_str(row.get("id")) or None
    title = cell_str(row.get("title"))
    if not title:
        raise ValueError("event missing title")
    status = cell_str(row.get("status")) or "confirmed"
    if status not in ("confirmed", "rain-delay", "canceled", "relocated"):
        status = "confirmed"
    doc = {
        "title": title,
        "host": cell_str(row.get("host")),
        "date": cell_str(row.get("date")),
        "time": cell_str(row.get("time")),
        "location": cell_str(row.get("location")),
        "description": cell_str(row.get("description")),
        "status": status,
        "statusNote": cell_str(row.get("statusNote")),
        "approved": cell_bool(row.get("approved")),
        "reminderSent": False,
        "createdBy": "import",
    }
    return eid, doc


def map_memory(row: dict) -> tuple[str, dict]:
    mid = cell_str(row.get("id")) or None
    title = cell_str(row.get("title"))
    if not title:
        raise ValueError("memory missing title")
    doc = {
        "title": title,
        "author": cell_str(row.get("author")),
        "date": cell_str(row.get("date")),
        "text": cell_str(row.get("text")),
        "approved": cell_bool(row.get("approved")),
        "createdBy": "import",
    }
    return mid, doc


def map_squad(row: dict) -> tuple[str, dict]:
    sid = cell_str(row.get("id")) or None
    name = cell_str(row.get("name"))
    if not name:
        raise ValueError("squad missing name")
    doc = {
        "name": name,
        "occupation": cell_str(row.get("occupation")),
        "age": cell_str(row.get("age")),
        "gender": cell_str(row.get("gender")),
        "socialLink": cell_str(row.get("socialLink")),
        "bio": cell_str(row.get("bio")),
        "photoUrl": cell_str(row.get("photoUrl")),
        "photoBase64": "",
        "photoMimeType": "image/jpeg",
        "approved": cell_bool(row.get("approved")),
        "createdBy": "import",
    }
    return sid, doc


def map_rsvp(row: dict) -> tuple[str, dict]:
    event_id = cell_str(row.get("eventId"))
    name = cell_str(row.get("name"))
    status = cell_str(row.get("status")) or "going"
    if status not in ("going", "not-going"):
        status = "going"
    if not event_id or not name:
        raise ValueError("rsvp missing eventId or name")
    # Legacy RSVPs are name-based; use a stable synthetic user id.
    user_id = f"legacy_{slug_name(name)}"
    rid = f"{user_id}_{event_id}"
    doc = {
        "eventId": event_id,
        "userId": user_id,
        "name": name,
        "status": status,
    }
    return rid, doc


MAPPERS = {
    "Events": ("events", map_event),
    "Memories": ("memories", map_memory),
    "Squad": ("squad", map_squad),
    "RSVPs": ("rsvps", map_rsvp),
}


def load_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    planned = []
    for sheet_name, (collection, mapper) in MAPPERS.items():
        if sheet_name not in wb.sheetnames:
            print(f"skip missing sheet: {sheet_name}")
            continue
        for row in rows_as_dicts(wb[sheet_name]):
            try:
                doc_id, doc = mapper(row)
            except ValueError as e:
                print(f"  skip {sheet_name} row: {e}")
                continue
            planned.append((collection, doc_id, doc))
    return planned


def get_firestore():
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        print("Install firebase-admin:  pip install firebase-admin", file=sys.stderr)
        sys.exit(1)

    if not firebase_admin._apps:
        raw = os.environ.get("FIREBASE_SERVICE_ACCOUNT_JSON", "").strip()
        if raw:
            info = json.loads(raw)
            cred = credentials.Certificate(info)
            firebase_admin.initialize_app(cred)
        elif os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
            firebase_admin.initialize_app()
        else:
            print(
                "Set FIREBASE_SERVICE_ACCOUNT_JSON or GOOGLE_APPLICATION_CREDENTIALS",
                file=sys.stderr,
            )
            sys.exit(1)
    return firestore.client()


def main():
    ap = argparse.ArgumentParser(description="Import spreadsheet → Firestore")
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be written without calling Firestore",
    )
    args = ap.parse_args()

    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    planned = load_workbook(args.xlsx)
    by_col = {}
    for col, doc_id, doc in planned:
        by_col.setdefault(col, 0)
        by_col[col] += 1
    print(f"Loaded {len(planned)} documents from {args.xlsx.name}:")
    for col, n in sorted(by_col.items()):
        print(f"  {col}: {n}")

    if args.dry_run:
        for col, doc_id, doc in planned[:8]:
            print(f"  [{col}] {doc_id or '(auto-id)'}: {json.dumps(doc, default=str)[:120]}…")
        if len(planned) > 8:
            print(f"  … and {len(planned) - 8} more")
        print("Dry run only — nothing written.")
        return

    db = get_firestore()
    batch = db.batch()
    batch_count = 0
    written = 0

    def commit():
        nonlocal batch, batch_count
        if batch_count:
            batch.commit()
            batch = db.batch()
            batch_count = 0

    for col, doc_id, doc in planned:
        ref = db.collection(col).document(doc_id) if doc_id else db.collection(col).document()
        # merge so re-runs update rather than fully clobber unknown fields
        batch.set(ref, doc, merge=True)
        batch_count += 1
        written += 1
        if batch_count >= 400:
            commit()

    commit()
    print(f"Wrote {written} documents to Firestore.")


if __name__ == "__main__":
    main()
